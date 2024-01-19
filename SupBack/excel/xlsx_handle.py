import os

import math

from openpyxl import Workbook, load_workbook

from win32com.client import Dispatch
import pythoncom

from log_handle import SupLogger


class ColumnNumHandler(object):
	def __init__(self):
		pass

	@staticmethod
	def single_chr_num(in_chr):
		return ord(in_chr) - ord("A") + 1

	@staticmethod
	def single_num_chr(in_num):
		return chr(in_num + ord("A"))

	def num_of_col(self, in_column):
		digit_len = len(in_column)
		num = 0
		for i in range(digit_len):
			if not in_column[i].isalpha():
				return
			num += self.single_chr_num(in_column[i]) * (26 ** (digit_len - i - 1))
		return num

	def col_of_num(self, in_num: int):
		if in_num < 1:
			SupLogger.info("invalid column num")
			return ""
		div, mod = divmod(in_num - 1, 26)
		return self.single_num_chr(div - 1) + self.single_num_chr(mod) if div else self.single_num_chr(mod)

	def form_col_list_by_range(self, col_range: [str]):
		assert len(col_range) == 2

		col_list = []

		col_min = self.num_of_col(col_range[0])
		col_max = self.num_of_col(col_range[1])

		column_nums = range(col_min, col_max + 1, 1)
		for column_num in column_nums:
			col_list.append(self.col_of_num(column_num))

		return col_list


class XlsxHandler(object):
	def __init__(self):
		self.file_path = ""
		self.wb = None
		self.ws = None
		self.sheet_index = 0
		self.sheet_index_win32com = 1

		self.wb_save = None
		self.ws_save = None

		self.attr_names = []
		self.attr_columns = []

		self.attrs = []

		self.col_handler = ColumnNumHandler()

		self.max_row = 0
		self.max_col = 0

	@staticmethod
	def is_valid_xlsx(in_file_path):
		return os.path.isfile(in_file_path) and os.path.exists(in_file_path) and in_file_path.endswith('.xlsx')

	def load(self, in_file_path, sheet_index=0):
		if not self.is_valid_xlsx(in_file_path):
			return

		self.file_path = in_file_path
		self.sheet_index = sheet_index
		self.sheet_index_win32com = sheet_index + 1
		self.wb = load_workbook(in_file_path, data_only=True)
		self.wb_save = load_workbook(in_file_path, data_only=False)

		ws_i = 0
		for sheet in self.wb:
			if ws_i == sheet_index:
				self.ws = sheet
				break
			ws_i += 1
		ws_i = 0
		for sheet in self.wb_save:
			if ws_i == sheet_index:
				self.ws_save = sheet
				break
			ws_i += 1

		self.max_row = self.ws.max_row
		self.max_col = self.ws.max_column

	def close(self):
		self.wb.close()
		self.wb_save.close()

	def get_cell_of_col_num(self, col_num: int, row_num: int):
		col = self.col_handler.col_of_num(col_num)
		cell_index = col + str(row_num)
		return self.ws[cell_index].value

	def get_cell_of_col(self, col: str, row_num: int):
		cell_index = col + str(row_num)
		return self.ws[cell_index].value

	def set_cell_of_col(self, col: str, row_num: int, value):
		cell_index = col + str(row_num)

		self.ws_save[cell_index] = value
			
	def get_attr_names_of_cols(self, col_list: [str], row_num):
		if self.ws is None:
			return []

		# 通过列名获取属性名列表
		attr_names = []
		for col in col_list:
			if col not in self.attr_columns:
				self.attr_columns.append(col)
			cell_index = col + str(row_num)
			attr_name = self.ws[cell_index].value
			if attr_name is None or attr_name == "":
				attr_name = col
			attr_names.append(attr_name)
		return attr_names

	def add_attr_name(self, col_list: [str], row_num: int):
		"""
		添加对应列行索引的表格的值，作为表头以及属性名
		:param col_list: 属性名表格的列名的列表（Excel本身列名，如["A","AA"]为一对起始）
		:param row_num: 属性名表格的行序号
		:return:
		"""
		if self.ws is None:
			return

		attr_names = self.get_attr_names_of_cols(col_list, row_num)
		for attr_name in attr_names:
			if attr_name not in self.attr_names:
				self.attr_names.append(attr_name)

	def add_attr_name_range(self, col_range: [str], row_num: int):
		"""
		添加对应列行索引的表格的值，作为表头以及属性名
		:param col_range: 属性名表格的列名的起始列表（Excel本身列名，如["A","AA"]为一对起始）
		:param row_num: 属性名表格的行序号
		:return:
		"""
		if self.ws is None:
			return

		col_list = self.col_handler.form_col_list_by_range(col_range)
		self.add_attr_name(col_list, row_num)

	def get_attr_col_by_name(self, attr_name):
		"""
		通过属性名获取属性列
		:param attr_name: 属性名（作为表头的表格的值，如“id”）
		:return: 属性列（Excel本身列名，如“A”,“AU”等）
		"""
		assert len(self.attr_names) == len(self.attr_columns)
		return self.attr_columns[self.attr_names.index(attr_name)]

	def get_attr_name_by_col(self, attr_col):
		"""
		通过属性列获取属性名
		:param attr_col: 属性列（Excel本身列名，如“A”,“AU”等）
		:return: 属性名（作为表头的表格的值，如“id”）
		"""
		assert len(self.attr_names) == len(self.attr_columns)
		return self.attr_names[self.attr_columns.index(attr_col)]

	def write_save_cell(self, col, row_num, value):
		pythoncom.CoInitialize()
		xlsx_app = Dispatch("Excel.Application")
		xlsx_app.Visible = False
		xlsx_wb = xlsx_app.Workbooks.Open(self.file_path)

		xlsx_wb.Worksheets(self.sheet_index_win32com).Range(col + str(row_num)).value = value

		xlsx_wb.Save()
		xlsx_wb.Close()
		xlsx_app.Quit()
		SupLogger.info("cell saved")

	def write_save_row_win32com(self, row_num, attr_dict):
		"""
		写入某行的属性字典并保存
		:param row_num: Excel自带的行序号（如1, 2）
		:param attr_dict: 属性字典，key为可以找到对应列序号的属性名， value为属性值
		:return:
		"""
		pythoncom.CoInitialize()
		xlsx_app = Dispatch("Excel.Application")
		xlsx_app.Visible = False
		xlsx_wb = xlsx_app.Workbooks.Open(self.file_path)

		for attr_name, attr_value in attr_dict.items():
			col = self.get_attr_col_by_name(attr_name)
			xlsx_wb.Worksheets(self.sheet_index_win32com).Range(col + str(row_num)).value = attr_value

		xlsx_wb.Save()
		xlsx_wb.Close()
		xlsx_app.Quit()

		SupLogger.info("excel saved")

	def delete_win32com(self, row_num):
		pythoncom.CoInitialize()
		xlsx_app = Dispatch("Excel.Application")
		xlsx_app.Visible = False
		xlsx_wb = xlsx_app.Workbooks.Open(self.file_path)

		xlsx_wb.Worksheets(self.sheet_index_win32com).Rows(row_num).Delete()

		xlsx_wb.Save()
		xlsx_wb.Close()
		xlsx_app.Quit()

		SupLogger.info(f"Deleted row: {row_num}")


if __name__ == "__main__":
	xlsx_handler = XlsxHandler()
	xlsx_handler.load(r"E:\Workflow\Block-wangjunyi.42-trunk\Client\Data\JungoTownRP\1_体素配置_RP.xlsx")
	xlsx_handler.add_attr_name_range(["C", "E"], 3)
	print(xlsx_handler.attr_names)
	column_handler = ColumnNumHandler()
	# print(column_handler.num_of_col("AZ"))
	print(column_handler.col_of_num(1))
	print(column_handler.col_of_num(26))
	print(column_handler.col_of_num(53))
	print(xlsx_handler.max_row)
